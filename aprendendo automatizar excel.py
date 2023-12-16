import openpyxl # blibioteca para abrir excel

# Sheets - sao paginas de uma planilha
# workbook - e o arquivo que contem estas paginas


planilha_test = openpyxl.Workbook()
planilha_test.create_sheet('Frutas') 
planilha_test.create_sheet('Legumes')
planilha_test.create_sheet('Sementes')
print(planilha_test.sheetnames)

cabecalho_frutas = ['titulos', 'Localizacao', 'Preco']
sheet_frutas = planilha_test.get_sheet_by_name('Frutas')
sheet_frutas.append(cabecalho_frutas)


frutas = [['Uva', 'Mercado', 5], ['Melancia', 'Mercado', 15], ['Bolo', 'Mercado', 40]]

for fruta in frutas:        #para adicionar a planilha
    sheet_frutas.append(fruta)

planilha_test.save('Dados supermecado.xlsx')  #sempre no final para salvar a planilha



#  // Comando para vizualizar ou editar planilha existente

# planilha = openpyxl.load_workbook('nome da planilha.xlsx')
# print(planila.sheetnames)  

#  // obtendo o sheet a trabalhar
# sheet1 = planilha.get_sheet_by_name('sheet1')
# sheet1['C3'].value = 'xxx'                                    #alter o valor do  C3 
# print(sheet1['C3'].value)                                     #imprimindo o valor na tela

#  // Percorrer uma planilha:

# for linha in sheet1.iter_rows(min_row=1,max_row=2, min_col=1, max_col=2):
#     print(linha[0].value,linha[1].value,linha[2].value)       #mostra os 3 colunas apenas, OBS: .value sempre no final

#  // para iterar apenas uma linha completa:

# for linha in sheet.iter_cols(min_col=2,max_col=2):            #tem o min_row=minimo(1,2,3,4,etc)
#     for cell in linha: 
#         print(cell.value)


